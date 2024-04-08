import tkinter as tk
from tkinter import ttk
import openpyxl
import os

class ExcelViewer:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel Viewer")
        self.root.state('zoomed')  # 窗口最大化

        self.current_sheet_index = 0
        self.workbook = None
        self.sheet_names = []

        self.load_workbook()
        self.create_widgets()
        self.show_sheet(self.sheet_names[0])  # 默认显示第一个表格

    def load_workbook(self):
        # 获取当前目录下的List.xlsx文件
        files = [f for f in os.listdir('.') if os.path.isfile(f)]
        excel_files = [f for f in files if f.endswith('.xlsx')]
        if not excel_files:
            raise FileNotFoundError("No Excel file found in the current directory.")

        self.workbook = openpyxl.load_workbook(excel_files[0])
        self.sheet_names = self.workbook.sheetnames

    def create_widgets(self):
        # 创建左侧按钮
        self.sheet_buttons_frame = ttk.Frame(self.root)
        self.sheet_buttons_frame.pack(side=tk.LEFT, fill=tk.Y)

        for sheet_name in self.sheet_names:
            button = ttk.Button(self.sheet_buttons_frame, text=sheet_name, command=lambda name=sheet_name: self.show_sheet(name))
            button.pack(fill=tk.X)

        # 创建treeview
        self.treeview_frame = ttk.Frame(self.root)
        self.treeview_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.treeview = ttk.Treeview(self.treeview_frame, columns=(),show='headings',style='Treeview')
        self.treeview.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        self.treeview_scroll = ttk.Scrollbar(self.treeview_frame, orient="vertical", command=self.treeview.yview)
        self.treeview_scroll.pack(side=tk.RIGHT, fill=tk.Y)
        self.treeview.configure(yscrollcommand=self.treeview_scroll.set)



    def show_sheet(self, sheet_name):
        # 清空treeview
        for i in self.treeview.get_children():
            self.treeview.delete(i)

        # 获取选择的工作表
        sheet = self.workbook[sheet_name]
        headers = []
        data = []

        # 读取表头
        for cell in sheet[1]:
            headers.append(cell.value)

        # 读取数据
        for row in sheet.iter_rows(min_row=2):
            # 检查整行是否为空
            if all(cell.value is None for cell in row):
                continue
            data_row = []
            for cell in row:
                data_row.append(cell.value if cell.value is not None else "")
            data.append(data_row)

        # 设置treeview的列和数据
        self.treeview['columns'] = headers
        print(headers)
        self.treeview.heading("#0", text="Index")
        self.treeview.column("#0", width=50)
        for header in headers:
            max_length = max(len(str(row[headers.index(header)])) for row in data)
            self.treeview.heading(header, text=header)
            self.treeview.column(header, width=max_length * 10, anchor=tk.CENTER)  # 设置列宽度为最长字符串长度的10倍

        # 插入数据
        for i, row in enumerate(data):
            self.treeview.insert("", 'end', text=str(i), values=row)

def main():
    root = tk.Tk()
    app = ExcelViewer(root)
    root.mainloop()

if __name__ == "__main__":
    main()

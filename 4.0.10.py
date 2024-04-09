import tkinter as tk
from tkinter import ttk
import openpyxl,sys,os,time
import tkinter.font as tkFont
from tkinter import *
from tkinter import messagebox
from collections import defaultdict as ded

class Tooltip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.widget.bind("<Enter>", self.show_tooltip)
        self.widget.bind("<Leave>", self.hide_tooltip)
    def show_tooltip(self, event):
        x, y, cx, cy = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 60
        y += self.widget.winfo_rooty() + 15
        self.tooltip = tk.Toplevel(self.widget)
        self.tooltip.overrideredirect(True)
        self.tooltip.geometry("+%d+%d" % (x, y))
        label = tk.Label(self.tooltip, text=self.text, justify="left",
                         background="#ffffe0", relief="solid", borderwidth=1,
                         font=("tahoma", "10", "normal"))
        label.pack(ipadx=1)
    def hide_tooltip(self, event):
        if self.tooltip:
            self.tooltip.destroy()
class ASSIST():
    def spass(self):
        pass
    def tableview_sortColumn(self,t):
        self.tlst=[(self.table.set(st,t),st)
                    for st in self.table.get_children("")]
        self.tlst.sort(reverse=self.reverseFlag)
        for index,item in enumerate(self.tlst):
            self.table.move(item[1],"",index)
        self.reverseFlag= not self.reverseFlag
    def pwsetclear(self):
        for widget in self.pwset.winfo_children():
            widget.destroy()
    def pwwindowclear(self):
        for widget in self.pwwindow.winfo_children():
            widget.destroy()
    def check_is_num(self,ss: str):
        try:
            float(ss)
        except:
            return False
        return True
    def delete_button(self, button):
        if messagebox.askyesno("确认", f"确认删除 {button.cget('''text''')}？"):
            if messagebox.askyesno("再次确认", f"确认删除 {button.cget('''text''')}？\n删除后无法恢复！"):
                button.destroy()
class Selector(tk.Toplevel):
    def __init__(self, parent, ded, update_callback):
        super().__init__(parent)
        self.title("Tag Selector")
        self.firstget = True
        self.slector_palce(parent)
        self.update_callback = update_callback
        self.selected_tags = set()
        self.selected_tags_label = tk.Label(self, text="Selected Tags:")
        self.selected_tags_label.pack()
        self.selected_tags_display = tk.Label(self, text="",wraplength=155,height=2)
        self.selected_tags_display.pack()
        self.search_label = tk.Label(self, text="Search:")
        self.search_label.pack()
        self.search_var = tk.StringVar()
        self.search_var.trace("w", self.filter_tags)
        self.search_entry = tk.Entry(self, textvariable=self.search_var)
        self.search_entry.pack()
        self.tag_frame = ttk.Frame(self)
        self.tag_frame.pack()
        self.tag_scrollbar = ttk.Scrollbar(self.tag_frame, orient=tk.VERTICAL)
        self.tag_scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.tag_canvas = tk.Canvas(self.tag_frame, yscrollcommand=self.tag_scrollbar.set)
        self.tag_canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.tag_scrollbar.config(command=self.tag_canvas.yview)
        self.tag_inner_frame = ttk.Frame(self.tag_canvas)
        self.tag_canvas.create_window((0, 0), window=self.tag_inner_frame, anchor=tk.NW)
        self.tag_checkboxes = []
        self.original_pack_options = {}
        for tag in ded:
            checkbox = ttk.Checkbutton(self.tag_inner_frame, text=tag+f" ({ded[tag]})", command=lambda t=tag: self.toggle_tag(t))
            checkbox.state(['!alternate'])
            self.tag_checkboxes.append(checkbox)
            self.original_pack_options[checkbox] = {"side": "top"}
            checkbox.pack(side="top", fill="x")
        self.filter_tags()
        self.tag_inner_frame.bind("<Configure>", self.on_frame_configure)
        self.lift(parent)
        self.resizable(False, False)
        self.protocol("WM_DELETE_WINDOW", self.hide_window)
    def hide_window(self):
        self.withdraw()
    def toggle_tag(self, tag):
        if tag in self.selected_tags:
            self.selected_tags.remove(tag)
        else:
            self.selected_tags.add(tag)
        self.update_selected_tags_display()
        self.update_callback(list(self.selected_tags))
    def toggle_tag(self, tag):
        if tag in self.selected_tags:
            self.selected_tags.remove(tag)
        else:
            self.selected_tags.add(tag)
        self.update_selected_tags_display()
        self.update_callback(list(self.selected_tags))
    def update_selected_tags_display(self):
        self.selected_tags_display.config(text=", ".join(sorted(self.selected_tags)))
    def filter_tags(self, *args):
        search_term = self.search_var.get().lower()
        for checkbox in self.tag_checkboxes:
            tag = checkbox.cget("text")
            if search_term in tag.lower():
                checkbox.pack(**self.original_pack_options[checkbox])
            else:
                checkbox.pack_forget()
    def on_frame_configure(self, event):
        self.tag_canvas.configure(scrollregion=self.tag_canvas.bbox("all"))
    def slector_palce(self,parent):
        self.geometry('160x330')
        parent_x, parent_y = parent.winfo_x(), parent.winfo_y()
        parent_width, parent_height = parent.winfo_width(), parent.winfo_height()
        toplevel_width, toplevel_height = self.winfo_width(), self.winfo_height()
        toplevel_x = parent_x + parent_width - toplevel_width
        toplevel_y = parent_y + parent_height - toplevel_height
        if self.firstget==True:
            toplevel_x-=160
            toplevel_y-=330
            self.firstget=False
        self.geometry(f'+{toplevel_x}+{toplevel_y}')
        self.lift()
    def clear_all(self):
        for checkbox in self.tag_checkboxes:
            checkbox.state(['!selected'])
        self.selected_tags = set()
        self.update_selected_tags_display()

class RangeSelector(tk.Toplevel):
    def __init__(self, parent, lowest, highest, rang, update_callback,des_callback):
        super().__init__(parent)
        self.title("Range Selector")
        self.firstget = True
        self.update_callback = update_callback
        self.lowest_value = tk.DoubleVar()
        self.highest_value = tk.DoubleVar()
        self.lowest_value.set(lowest)
        self.highest_value.set(highest)
        self.lowest = lowest
        self.highest = highest
        self.des_callback = des_callback
        self.config(bg='white')

        tk.Label(self, text="最低值",bg='white').grid(row=0, column=0,ipadx=10,ipady=10)
        self.lowest_entry = tk.Entry(self, textvariable=self.lowest_value, width=6)
        self.lowest_entry.grid(row=0, column=1,padx = 20)
        self.lowest_slider = tk.Scale(self ,bg='white',from_=lowest, to=highest-0.1, resolution=rang, variable=self.lowest_value, orient=tk.HORIZONTAL, command=self.update_highest_min)
        self.lowest_slider.grid(row=1, column=0, columnspan=2, ipady=5)

        tk.Label(self, text="最高值",bg='white').grid(row=2, column=0,ipadx=10,ipady=10)
        self.highest_entry = tk.Entry(self, textvariable=self.highest_value, width=6)
        self.highest_entry.grid(row=2, column=1,padx = 20)
        self.highest_slider = tk.Scale(self,bg='white', from_=lowest+0.1, to=highest, resolution=rang, variable=self.highest_value, orient=tk.HORIZONTAL, command=self.update_lowest_max)
        self.highest_slider.grid(row=3, column=0, columnspan=2, ipady=5)

        self.place_window(parent)
        self.protocol("WM_DELETE_WINDOW", self.des_callback)

    def update_highest_min(self, value):
        self.highest_slider.config(from_=max(float(value)+0.1, self.lowest))
        self.filter_table()

    def update_lowest_max(self, value):
        self.lowest_slider.config(to=min(self.highest, float(value)-0.1))
        self.filter_table()

    def filter_table(self):
        self.update_callback(self.lowest_value.get(), self.highest_value.get())

    def place_window(self, parent):
        self.geometry('160x200')
        parent_x, parent_y = parent.winfo_x(), parent.winfo_y()
        parent_width, parent_height = parent.winfo_width(), parent.winfo_height()
        toplevel_width, toplevel_height = self.winfo_width(), self.winfo_height()
        toplevel_x = parent_x + parent_width - toplevel_width
        toplevel_y = parent_y + parent_height - toplevel_height
        if self.firstget:
            toplevel_x -= 160
            toplevel_y -= 200
            self.firstget = False
        self.geometry(f'+{toplevel_x}+{toplevel_y}')
        self.lift()


class Gui(Tooltip,ASSIST,Selector):
    def __init__(self):
        self.root = tk.Tk()
        self.initfont()
        self.picload()
        self.initwindowsize()
        self.initwindowpaned()
        self.sheetchoice()
        self.set()
        self.initmain()
        self.load_workbook()
        self.create_sheetbtn()

        self.show_sheet(self.sheet_names[0])

        self.root.mainloop()

    '''初始化界面设置'''
    def initwindowsize(self):
        #设定最初始的界面大小
        self.root.title('My Animate list')
        self.root.geometry('1680x530')
        self.new_width = 1680
        self.new_height = 530
        self.root.minsize(1490,530)
        self.rootlaststate='normal'
        self.root.configure(bg='ghostwhite')
        self.root.option_add("*FontQuality", "high")
        self.root.bind("<Configure>", self.on_zoom)
    def on_zoom(self,event):
        #完成是否全屏界面的切换
        self.new_width = self.root.winfo_width()
        self.new_height = self.root.winfo_height()
        if self.root.state() == "zoomed" and self.rootlaststate=='normal':
            self.root.resizable(width=True, height=True)
            self.root.geometry('2160x1440')
            self.root.state('zoomed')
            self.rootlaststate= "zoomed"
        elif self.root.state() == "normal" and self.rootlaststate=='zoomed':
            self.root.geometry('1680x540')
            self.root.resizable(width=False, height=True)
            self.rootlaststate="normal"
        self.root.rowconfigure(0, weight=1)
        for child in self.root.winfo_children():
            if isinstance(child, Selector):
                child.slector_palce(self.root)
            elif isinstance(child,RangeSelector):
                child.place_window(self.root)
    def initfont(self):
        self.standard_font = tkFont.Font(family="等线", size=11, weight="normal")
        self.special_font = tkFont.Font(family="等线", size=11, weight="bold")

    def initwindowpaned(self):
        self.pwsheet = tk.PanedWindow(self.root,width=60,orient= "vertical",bg='white')
        self.pwset = tk.PanedWindow(self.root,width=60,bg='white')
        self.pwwindow  = tk.PanedWindow(self.root,width = self.new_width-120,bg='white')
        self.pwsheet.grid(row=0,column=0,ipadx=2,sticky='nwes')
        self.pwset.grid(row=0,column=1,sticky='nwes')
        self.pwwindow.grid(row=0,column=2,ipadx=2,sticky='nwes')
        #self.pwset.grid_remove()


    '''初始化表选择'''
    def sheetchoice(self):
        self.sheet_buttons_frame = ttk.Frame(
            master = self.pwsheet,
            width = 80)
        self.sheet_buttons_frame.grid(row=0, column=0, sticky=tk.N+tk.S)
    '''初始化图片'''
    def picload(self):
        self.path = getattr(sys, '_MEIPASS', '.')
        self.bt0png0 = PhotoImage(file=self.path +'/outicon0.png')
        self.bt0png1 = PhotoImage(file=self.path +'/outicon1.png')
        #self.bt1png0 = PhotoImage(file=r'icons/mainicon0.png')
        self.bt1png1 = PhotoImage(file=self.path +'/mainicon1.png')
        self.bt2png0 = PhotoImage(file=self.path +'/igicon0.png')
        self.bt2png1 = PhotoImage(file=self.path +'/igicon1.png')
        self.bt3png0 = PhotoImage(file=self.path +'/ieicon0.png')
        self.bt3png1 = PhotoImage(file=self.path +'/ieicon1.png')
        self.addpng = PhotoImage(file=self.path+'/add.png')
        self.deletepng = PhotoImage(file=self.path+'/delete.png')
        self.editpng = PhotoImage (file=self.path +'/edit.png')
        self.tagspng = PhotoImage(file=self.path+'/tags.png')
        self.coppng = PhotoImage(file =self.path+'/cop.png')
        self.markpng = PhotoImage(file =self.path+'/mark.png')
        self.recpng = PhotoImage(file =self.path+'/rec.png')
        self.savepng =PhotoImage(file = self.path+'/save.png')
        self.showallpng =PhotoImage(file = self.path+'/showall.png')
        self.haveseenpng=PhotoImage(file = self.path+'/haveseen.png')
        self.startseenpng=PhotoImage(file = self.path+'/startseen.png')
    '''初始化添加删改栏目'''
    def set(self):
        self.pwsetclear()
        self.btadd= tk.Button(self.pwset,height= 30,width=30,image=self.addpng,
            bg='white',relief=GROOVE,activebackground='lime',command =self.spass)
        self.btdelete= tk.Button(self.pwset,height= 30,width=30,image=self.deletepng,
            bg='white',relief=GROOVE,activebackground='crimson',command =self.spass)
        self.btedit = tk.Button(self.pwset,height =30,width = 30,image=self.editpng,
            bg='white',relief=GROOVE,activebackground='teal',command=self.settableeditable)
        self.btsave = tk.Button(self.pwset,height =30,width = 30,image=self.savepng,
            bg='white',relief=GROOVE,activebackground='yellow',command=self.spass)
        self.sep1 =ttk.Separator(self.pwset, orient="horizontal")
        self.bttags = tk.Button(self.pwset,height =30,width = 30,image=self.tagspng,
            bg='white',relief=GROOVE,activebackground='violet',command= self.settabletagf)
        self.btcops=tk.Button(self.pwset,height =30,width = 30,image=self.coppng,
            bg='white',relief=GROOVE,activebackground='violet',command= self.settablecf)
        self.btmark=tk.Button(self.pwset,height =30,width = 30,image=self.markpng,
            bg='white',relief=GROOVE,activebackground='violet',command= self.settableM)
        self.btrec=tk.Button(self.pwset,height =30,width = 30,image=self.recpng,
            bg='white',relief=GROOVE,activebackground='violet',command= self.settableR)
        self.sep2 =ttk.Separator(self.pwset, orient="horizontal")
        self.btshowall=tk.Button(self.pwset,height =30,width = 30,image=self.showallpng,
            bg='white',relief=GROOVE,activebackground='orange',command= self.showall)

        Tooltip(self.btadd,'Add')
        Tooltip(self.btdelete,'Delete')
        Tooltip(self.btedit,'Edit')
        Tooltip(self.btsave,'Save')
        Tooltip(self.bttags,'Filter-Tags')
        Tooltip(self.btcops,'Filter-Co.')
        Tooltip(self.btmark,'Filter-Mark.')
        Tooltip(self.btrec,'Filter-Recommendation')
        Tooltip(self.btshowall,'Show-all')

        self.btadd.grid(row=0,ipadx=10,ipady=10)
        self.btdelete.grid(row=1,ipadx=10,ipady=10)
        self.btedit.grid(row=2,ipadx=10,ipady=10)
        self.btsave.grid(row=3,ipadx=10,ipady=10)
        self.sep1.grid(row=4,sticky='ew',padx=5,ipadx=10,pady=6)
        self.bttags.grid(row=5,ipadx=10,ipady=10)
        self.btcops.grid(row=6,ipadx=10,ipady=10)
        self.btmark.grid(row=7,ipadx=10,ipady=10)
        self.btrec.grid(row=8,ipadx=10,ipady=10)
        self.sep2.grid(row=9,sticky='ew',padx=5,ipadx=10,pady=4)
        self.btshowall.grid(row=10,ipadx=10,ipady=10)
    '''初始化表'''
    def initmain(self):
        self.tecols=['中文名', '原名', '评分', '画面', '剧本', '配乐', '推荐度', '制作方', '发行时间', '剧话信息', 'TAG']
        self.widths=[300,       325,    50,     35,     35,     35,     50,     175,        70,     175,        225]
        self.minwidths=[300,    50,     50,     35,     35,     35,     50,     0,         70,     50,         0]
        self.cansort=[False,    False,  True,   True,   True,   True,   True,   True,       True,   False,      False]
        self.table_edit_entry= None
        self.original_values = {}
        self.editing_item = None
        self.editing_column = None
        self.tableeditable = False
        self.alltag = ded(lambda : 0)
        self.allc = ded(lambda: 0)
        self.selected_T = None
        self.selected_C = None
        self.btedit.config(bg = 'white')
        self.yscroll = tk.Scrollbar(self.pwwindow,width =5)
        self.yscroll.pack(fill='y', expand=True)
        self.table = ttk.Treeview(
            master=self.pwwindow,
            columns=self.tecols,
            show='headings',
            style='Treeview',
            #xscrollcommand=self.xscroll.set,
            yscrollcommand=self.yscroll.set
        )
        self.yscroll.config(command=self.table.yview)
        self.reverseFlag=False
        for i in range(11):
            self.table.heading(column=self.tecols[i],text=self.tecols[i],anchor='center',command =(lambda t=self.tecols[i]: self.tableview_sortColumn(t)) if self.cansort[i] else self.spass)
            self.table.column(self.tecols[i],width = self.widths[i],minwidth = self.minwidths[i],stretch=True)
        self.pwwindow.add(self.table)
        self.pwwindow.add(self.yscroll)
        self.hidden_rows= set()
        self.filter_using = None
        self.range_using = None
        self.lowM=6.0
        self.highM=10.0
        self.lowR=0.0
        self.highM=6.0

    '''单击右键编辑'''
    def table_edit(self,event):
        if self.table_edit_entry is not None:
            return
        item = self.table.identify_row(event.y)
        column = self.table.identify_column(event.x)
        if item and column != "#0":
            bbox = self.table.bbox(item, column)
            if bbox:
                x, y, width, height = bbox
                column_index = int(column.split("#")[-1]) - 1
                original_value = self.table.item(item, "values")[column_index]
                self.original_values[(item, column)] = original_value
                self.table_edit_entry = tk.Entry(self.table)
                self.table_edit_entry.place(x=x, y=y, width=width, height=height)
                self.table_edit_entry.insert(0, original_value)
                self.table_edit_entry.focus_set()
                self.table_edit_entry.bind("<Return>", lambda event, item=item, column=column: self.table_update_cell(event, item, column))
                self.editing_item = item
                self.editing_column = column

    def table_update_cell(self, event, item, column):
        if self.table_edit_entry is None:
            return
        new_value = self.table_edit_entry.get()
        original_value = self.original_values.get((item, column), "")
        row_name = self.table.item(item, "values")[0]
        column_name = self.table.heading(column, option="text")
        confirmation = messagebox.askyesno("数据更改", f"{row_name} 的 {column_name} 将从 {original_value} 更改为 {new_value}?")
        if confirmation:
            if self.check(column_name,new_value):
                values = list(self.table.item(item, "values"))
                column_index = int(column.split("#")[-1]) - 1
                values[column_index] = new_value
                self.table.item(item, values=values)
            else:
                messagebox.showerror('数据错误',f"{new_value} 不符合 {column_name}列填写规则！")
                return
        self.table_edit_entry.destroy()
        self.table_edit_entry = None
        del self.original_values[(item, column)]


    def table_cancel_edit(self, event):
        if self.table_edit_entry is None:
            return
        if self.table_edit_entry.get() == self.original_values.get((self.editing_item, self.editing_column), ""):
            self.table_edit_entry.destroy()
            self.table_edit_entry = None
            del self.original_values[(self.editing_item, self.editing_column)]
        else:
            self.table_update_cell(event, self.editing_item, self.editing_column)

    '''读取数据'''
    def load_workbook(self):
        files = [f for f in os.listdir('.') if os.path.isfile(f)]
        excel_files = [f for f in files if f.endswith('.xlsx')]
        if not excel_files:
            raise FileNotFoundError("No Excel file found in the current directory.")
        self.workbook = openpyxl.load_workbook(excel_files[0])
        self.sheet_names = self.workbook.sheetnames
    '''根据数据得到表名'''
    def create_sheetbtn(self):
        i=0

        for sheet_name in self.sheet_names:
            button = tk.Button(master=self.pwsheet,
                bg='white',
                width = 15,
                pady = 10,
                text=sheet_name,
                font=self.standard_font,
                relief=GROOVE,
                activebackground='#E2EFDA',
                command=lambda name=sheet_name: self.show_sheet(name))
            button.bind("<Button-3>", lambda event, b=button: self.sheetbtn_menu(event, b))
            button.grid(row = i)
            i+=1
    '''表点击时颜色更改'''
    def sheetbtn_color(self, sheet_name):
        for child in self.pwsheet.winfo_children():
            if isinstance(child, tk.Button):
                if child.cget("text")==sheet_name:
                    child.config(bg='#70AD47')
                else:
                    child.config(bg='white')

    '''表右键'''
    def sheetbtn_menu(self, event, button):
        menu = tk.Menu(self.pwsheet, tearoff=0, font=('方正新书宋_GBK', 10, 'bold'))
        menu.add_command(label="锁定", command=lambda: None,background='white', activebackground='#E2EFDA', activeforeground='black')
        menu.add_command(label="删除", command=lambda: self.delete_button(button),background='white', activebackground='#E2EFDA',activeforeground='red')
        menu.post(button.winfo_rootx() + button.winfo_width(), button.winfo_rooty())
    '''显示表'''
    def show_sheet(self, sheet_name):
        self.pwwindowclear()
        self.initmain()
        self.sheetbtn_color(sheet_name)
        for i in self.table.get_children():
            self.table.delete(i)
        sheet = self.workbook[sheet_name]
        headers = []
        data = []
        for cell in sheet[1]:
            headers.append(cell.value)
        for row in sheet.iter_rows(min_row=2):
            if all(cell.value is None for cell in row):
                continue
            data_row = []
            for cell in row:
                data_row.append(cell.value if cell.value is not None else "")
            data.append(data_row)

        for i, row in enumerate(data):
            thisrow = []
            t= 0
            for j in self.tecols:
                if j in headers:
                    thisrow.append(row[t])
                    t+=1
                else:
                    thisrow.append("")
            self.table.insert("", 'end', text=str(i), values=thisrow)

        for col in self.table['columns']:
            if col not in headers:
                self.table.column(col, width=0, stretch=False)
        self.tagget()
        self.cget()
        for child in self.root.winfo_children():
            if isinstance(child, Selector):
                child.destroy()
                break
    def tagget(self):
        for item in self.table.get_children():
            _tags = self.table.item(item)['values'][10].split()  # Assuming 'TAG' is at index 10
            for _tag in _tags:
                self.alltag[_tag]+=1
        #print(self.alltag)
    def cget(self):
        for item in self.table.get_children():
            _cs = self.table.item(item)['values'][7].split('&')  # Assuming 'TAG' is at index 10
            for _c in _cs:
                self.allc[_c.strip()]+=1
        #print(self.allc)

    def settableeditable(self):
        if self.tableeditable == False :
            messagebox.showinfo("编辑中","编辑已打开!")
            self.btedit.config(bg = '#E2EFDA')
            self.table.bind("<Button-3>", self.table_edit)
            self.table.bind("<Button-1>", self.table_cancel_edit)
            self.tableeditable = True
        else:
            if self.table_edit_entry is not None:
                messagebox.showerror("错误", "编辑进行中！")
                return
            messagebox.showinfo("编辑关闭","编辑已关闭!")
            self.btedit.config(bg = 'white')
            self.table.unbind("<Button-3>")
            self.table.unbind("<Button-1>")
            self.tableeditable = False
    def check(self,column_name,new_value):
        self.checkcolumn=[ '评分', '画面', '剧本', '配乐', '推荐度']
        if column_name in self.checkcolumn:
            if column_name in ['评分','推荐度']:
                if self.check_is_num(new_value) and 5<=float(new_value)<=10:
                    return True
                else:
                    return False
            else:
                if new_value.strip() in ['S+','S','A+','A','B','C','D','E']:
                    return True
                else:
                    return False
        else:
            return True
    def settabletagf(self):
        for child in self.root.winfo_children():
            if isinstance(child, RangeSelector):
                child.destroy()
                self.tablereset()
                break
        if self.filter_using == None:
            Selector(self.root,self.alltag,self.update_selected_t)
            self.filter_using = 'TAG'
        elif self.filter_using =='TAG':
            for child in self.root.winfo_children():
                if isinstance(child, Selector):
                    child.deiconify()
                    return
        else:
            self.showall()
            for child in self.root.winfo_children():
                if isinstance(child, Selector):
                    #child.deiconify()
                    child.destroy()
                    break
            Selector(self.root,self.alltag,self.update_selected_t)
            self.filter_using = 'TAG'

    def settableM(self):
        for child in self.root.winfo_children():
            if isinstance(child, Selector):
                child.destroy()
                self.tablereset()
                self.filter_using =None
                break

        if self.range_using == None:
            RangeSelector(self.root,6.0,10.0,0.1,self.update_mark,self.des_callback)
            self.range_using = 'M'
        elif self.range_using == 'M':
            return
        else:
            self.des_callback()
            RangeSelector(self.root,6.0,10.0,0.1,self.update_mark,self.des_callback)
            self.range_using = 'M'

    def settableR(self):
        for child in self.root.winfo_children():
            if isinstance(child, Selector):
                child.destroy()
                self.tablereset()
                self.filter_using =None
                break
        #print(self.range_using)
        if self.range_using == None:
            RangeSelector(self.root,0.0,6.0,0.5,self.update_recom,self.des_callback)
            self.range_using = 'R'
        elif self.range_using == 'R':
            return
        else:
            self.des_callback()
            RangeSelector(self.root,0.0,6.0,0.5,self.update_recom,self.des_callback)
            self.range_using = 'R'



    def update_selected_t(self,tags):
        self.selected_T=tags
        self.tablereset()
        if not self.selected_T:
            return
        for item in self.table.get_children():
            tags_of_item = self.table.item(item)['values'][10].split()
            if any(tag in tags_of_item for tag in self.selected_T):
                continue
            else:
                self.table.detach(item)
                self.hidden_rows.add(item)
        self.sortshow()
    def tablereset(self):
        for row_id in self.hidden_rows:
            self.table.reattach(row_id, '', 'end')
        self.hidden_rows.clear()
        self.sortshow()

    def showall(self):
        self.tablereset()
        for child in self.root.winfo_children():
            if isinstance(child, Selector):
                child.clear_all()
            elif isinstance(child,RangeSelector):
                child.destory()

    def sortshow(self):
        self.tableview_sortColumn('评分')
        if self.reverseFlag:
            self.tableview_sortColumn('评分')

    def settablecf(self):
        for child in self.root.winfo_children():
            if isinstance(child, RangeSelector):
                child.destroy()
                self.tablereset()
                self.filter_using =None
                break
        if self.filter_using == None:
            Selector(self.root,self.allc,self.update_selected_c)
            self.filter_using = 'PDC'
        elif self.filter_using=='PDC':
            for child in self.root.winfo_children():
                if isinstance(child, Selector):
                    child.deiconify()
                    return
        else:
            self.showall()
            for child in self.root.winfo_children():
                if isinstance(child, Selector):
                    #child.deiconify()
                    child.destroy()
                    break
            Selector(self.root,self.allc,self.update_selected_c)
            self.filter_using = 'PDC'

    def update_selected_c(self,tags):
        for child in self.root.winfo_children():
            if isinstance(child, RangeSelector):
                child.destroy()
                self.tablereset()
                self.filter_using =None
                break
        self.selected_C=tags
        self.tablereset()
        if not self.selected_C:
            return
        for item in self.table.get_children():
            c_of_item = self.table.item(item)['values'][7].split('&')
            if any(c.strip() in c_of_item for c in self.selected_C):
                continue
            else:
                self.table.detach(item)
                self.hidden_rows.add(item)
        self.sortshow()

    def update_mark(self,low,hight):
        self.tablereset()
        self.lowM,self.highM=low,hight
        for item in self.table.get_children():
            thisM = float(self.table.item(item)['values'][2])
            if not self.lowM<=thisM<=self.highM:
                self.table.detach(item)
                self.hidden_rows.add(item)
        self.sortshow()

    def update_recom(self,low,hight):
        self.tablereset()
        self.lowM,self.highM=low,hight
        for item in self.table.get_children():
            thisM = float(self.table.item(item)['values'][6])
            if not self.lowM<=thisM<=self.highM:
                self.table.detach(item)
                self.hidden_rows.add(item)
        self.sortshow()

    def des_callback(self):
        for child in self.root.winfo_children():
            if isinstance(child, RangeSelector):
                child.destroy()
        self.tablereset()
        self.range_using == None



Gui()
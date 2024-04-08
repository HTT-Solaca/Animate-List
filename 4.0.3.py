import tkinter as tk
from tkinter import ttk
import openpyxl,sys,os
from tkinter import *
import tkinter.font as tkFont

class Tooltip:
    def __init__(self, widget, text):
        self.widget = widget
        self.text = text
        self.widget.bind("<Enter>", self.show_tooltip)
        self.widget.bind("<Leave>", self.hide_tooltip)

    def show_tooltip(self, event):
        x, y, cx, cy = self.widget.bbox("insert")
        x += self.widget.winfo_rootx() + 60
        y += self.widget.winfo_rooty() + 15
        self.tooltip = tk.Toplevel(self.widget)
        self.tooltip.overrideredirect(True)
        self.tooltip.geometry("+%d+%d" % (x, y))
        label = tk.Label(self.tooltip, text=self.text, justify="left",
                         background="#ffffe0", relief="solid", borderwidth=1,
                         font=("tahoma", "10", "normal"))
        label.pack(ipadx=1)

    def hide_tooltip(self, event):
        if self.tooltip:
            self.tooltip.destroy()

class ASSIST():
    def tableview_sortColumn(self,t):
        self.tlst=[(self.table.set(st,t),st)
                    for st in self.table.get_children("")]
        self.tlst.sort(reverse=self.reverseFlag)
        for index,item in enumerate(self.tlst):
            self.table.move(item[1],"",index)
        self.reverseFlag= not self.reverseFlag


    def pwsetclear(self):
        for widget in self.pwset.winfo_children():
            widget.destroy()
    def pwwindowclear(self):
        for widget in self.pwwindow.winfo_children():
            widget.destroy()

    def check_is_num(self,ss: str):
        try:
            float(ss)
        except:
            return False
        return True



class Gui(Tooltip,ASSIST):
    def __init__(self):
        self.root = tk.Tk()
        self.initfont()
        self.picload()
        self.initwindowsize()
        self.initwindowpaned()
        self.sheetchoice()
        self.set()
        self.initmain()
        self.load_workbook()
        self.create_widgets()
        self.show_sheet(self.sheet_names[0])

        self.root.mainloop()

    '''初始化界面设置'''
    def initwindowsize(self):
        #设定最初始的界面大小
        self.root.title('My Animate list')
        self.root.geometry('1680x530')
        self.new_width = 1680
        self.new_height = 530
        self.root.minsize(1490,530)
        self.rootlaststate='normal'
        self.root.configure(bg='ghostwhite')
        self.root.bind("<Configure>", self.on_zoom)
    def on_zoom(self,event):
        #完成是否全屏界面的切换
        self.new_width = self.root.winfo_width()
        self.new_height = self.root.winfo_height()
        if self.root.state() == "zoomed" and self.rootlaststate=='normal':
            self.root.resizable(width=True, height=True)
            self.root.geometry('2160x1440')
            self.root.state('zoomed')
            self.rootlaststate= "zoomed"
        elif self.root.state() == "normal" and self.rootlaststate=='zoomed':
            self.root.geometry('1680x540')
            self.root.resizable(width=False, height=True)
            self.rootlaststate="normal"
        self.root.rowconfigure(0, weight=1)
    def initfont(self):
        self.standard_font = tkFont.Font(family="等线", size=11, weight="normal")
        self.special_font = tkFont.Font(family="等线", size=11, weight="bold")

    def initwindowpaned(self):
        self.pwsheet = tk.PanedWindow(self.root,width=60,orient= "vertical",bg='white')
        self.pwset = tk.PanedWindow(self.root,width=60,bg='white')
        self.pwwindow  = tk.PanedWindow(self.root,width = self.new_width-120,bg='white')
        self.pwsheet.grid(row=0,column=0,ipadx=2,sticky='nwes')
        self.pwset.grid(row=0,column=1,sticky='nwes')
        self.pwwindow.grid(row=0,column=2,ipadx=2,sticky='nwes')
        #self.pwset.grid_remove()


    '''初始化表选择'''
    def sheetchoice(self):
        self.sheet_buttons_frame = ttk.Frame(
            master = self.pwsheet,
            width = 80)
        self.sheet_buttons_frame.grid(row=0, column=0, sticky=tk.N+tk.S)
    '''初始化图片'''
    def picload(self):
        self.path = getattr(sys, '_MEIPASS', '.')
        self.bt0png0 = PhotoImage(file=self.path +'/outicon0.png')
        self.bt0png1 = PhotoImage(file=self.path +'/outicon1.png')
        #self.bt1png0 = PhotoImage(file=r'icons/mainicon0.png')
        self.bt1png1 = PhotoImage(file=self.path +'/mainicon1.png')
        self.bt2png0 = PhotoImage(file=self.path +'/igicon0.png')
        self.bt2png1 = PhotoImage(file=self.path +'/igicon1.png')
        self.bt3png0 = PhotoImage(file=self.path +'/ieicon0.png')
        self.bt3png1 = PhotoImage(file=self.path +'/ieicon1.png')
        self.addpng = PhotoImage(file=self.path+'/add.png')
        self.deletepng = PhotoImage(file=self.path+'/delete.png')
        self.editpng = PhotoImage (file=self.path +'/edit.png')
        self.tagspng = PhotoImage(file=self.path+'/tags.png')
        self.coppng = PhotoImage(file =self.path+'/cop.png')
        self.markpng = PhotoImage(file =self.path+'/mark.png')
        self.recpng = PhotoImage(file =self.path+'/rec.png')
        self.savepng =PhotoImage(file = self.path+'/save.png')
        self.showallpng =PhotoImage(file = self.path+'/showall.png')
        self.haveseenpng=PhotoImage(file = self.path+'/haveseen.png')
        self.startseenpng=PhotoImage(file = self.path+'/startseen.png')
    '''初始化添加删改栏目'''
    def set(self):
        self.pwsetclear()
        self.btadd= tk.Button(self.pwset,height= 30,width=30,image=self.addpng,
            bg='white',relief=GROOVE,activebackground='lime',command =self.spass)
        self.btdelete= tk.Button(self.pwset,height= 30,width=30,image=self.deletepng,
            bg='white',relief=GROOVE,activebackground='crimson',command =self.spass)
        self.btedit = tk.Button(self.pwset,height =30,width = 30,image=self.editpng,
            bg='white',relief=GROOVE,activebackground='teal',command=self.spass)
        self.btsave = tk.Button(self.pwset,height =30,width = 30,image=self.savepng,
            bg='white',relief=GROOVE,activebackground='yellow',command=self.spass)
        self.sep1 =ttk.Separator(self.pwset, orient="horizontal")
        self.bttags = tk.Button(self.pwset,height =30,width = 30,image=self.tagspng,
            bg='white',relief=GROOVE,activebackground='violet',command= self.spass)
        self.btcops=tk.Button(self.pwset,height =30,width = 30,image=self.coppng,
            bg='white',relief=GROOVE,activebackground='violet',command= self.spass)
        self.btmark=tk.Button(self.pwset,height =30,width = 30,image=self.markpng,
            bg='white',relief=GROOVE,activebackground='violet',command= self.spass)
        self.btrec=tk.Button(self.pwset,height =30,width = 30,image=self.recpng,
            bg='white',relief=GROOVE,activebackground='violet',command= self.spass)
        self.sep2 =ttk.Separator(self.pwset, orient="horizontal")
        self.btshowall=tk.Button(self.pwset,height =30,width = 30,image=self.showallpng,
            bg='white',relief=GROOVE,activebackground='orange',command= self.spass)

        Tooltip(self.btadd,'Add')
        Tooltip(self.btdelete,'Delete')
        Tooltip(self.btedit,'Edit')
        Tooltip(self.btsave,'Save')
        Tooltip(self.bttags,'Filter-Tags')
        Tooltip(self.btcops,'Filter-Co.')
        Tooltip(self.btmark,'Filter-Mark.')
        Tooltip(self.btrec,'Filter-Recommendation')
        Tooltip(self.btshowall,'Show-all')

        self.btadd.grid(row=0,ipadx=10,ipady=10)
        self.btdelete.grid(row=1,ipadx=10,ipady=10)
        self.btedit.grid(row=2,ipadx=10,ipady=10)
        self.btsave.grid(row=3,ipadx=10,ipady=10)
        self.sep1.grid(row=4,sticky='ew',padx=5,ipadx=10,pady=6)
        self.bttags.grid(row=5,ipadx=10,ipady=10)
        self.btcops.grid(row=6,ipadx=10,ipady=10)
        self.btmark.grid(row=7,ipadx=10,ipady=10)
        self.btrec.grid(row=8,ipadx=10,ipady=10)
        self.sep2.grid(row=9,sticky='ew',padx=5,ipadx=10,pady=4)
        self.btshowall.grid(row=10,ipadx=10,ipady=10)
    def initmain(self):
        self.tecols=['中文名', '原名', '评分', '画面', '剧本', '配乐', '推荐度', '制作方', '发行时间', '剧话信息', 'TAG']
        self.widths=[300,       325,    50,     35,     35,     35,     50,     175,        70,     175,        225]
        self.minwidths=[300,    50,     50,     35,     35,     35,     50,     0,         70,     50,         0]
        self.cansort=[False,    False,  True,   True,   True,   True,   True,   True,       True,   False,      False]

        self.yscroll = tk.Scrollbar(self.pwwindow,width =5)
        self.yscroll.pack(fill='y', expand=True)
        self.table = ttk.Treeview(
            master=self.pwwindow,
            columns=self.tecols,
            show='headings',
            style='Treeview',
            #xscrollcommand=self.xscroll.set,
            yscrollcommand=self.yscroll.set
        )
        self.yscroll.config(command=self.table.yview)
        self.reverseFlag=False
        for i in range(11):
            self.table.heading(column=self.tecols[i],text=self.tecols[i],anchor='center',command =(lambda t=self.tecols[i]: self.tableview_sortColumn(t)) if self.cansort[i] else self.spass)
            self.table.column(self.tecols[i],width = self.widths[i],minwidth = self.minwidths[i],stretch=True)
        self.pwwindow.add(self.table)
        self.pwwindow.add(self.yscroll)

    def load_workbook(self):
        files = [f for f in os.listdir('.') if os.path.isfile(f)]
        excel_files = [f for f in files if f.endswith('.xlsx')]
        if not excel_files:
            raise FileNotFoundError("No Excel file found in the current directory.")
        self.workbook = openpyxl.load_workbook(excel_files[0])
        self.sheet_names = self.workbook.sheetnames
    def create_widgets(self):
        i=0
        for sheet_name in self.sheet_names:
            button = tk.Button(master=self.pwsheet,
                bg='white',
                width = 15,
                pady = 10,
                text=sheet_name,
                font=self.standard_font,
                relief=GROOVE,
                command=lambda name=sheet_name: self.show_sheet(name))
            button.grid(row = i)
            i+=1
    def show_sheet(self, sheet_name):
        self.pwwindowclear()
        self.initmain()
        for i in self.table.get_children():
            self.table.delete(i)
        sheet = self.workbook[sheet_name]
        headers = []
        data = []
        for cell in sheet[1]:
            headers.append(cell.value)
        for row in sheet.iter_rows(min_row=2):
            if all(cell.value is None for cell in row):
                continue
            data_row = []
            for cell in row:
                data_row.append(cell.value if cell.value is not None else "")
            data.append(data_row)

        for i, row in enumerate(data):
            thisrow = []
            t= 0
            for j in self.tecols:
                if j in headers:
                    thisrow.append(row[t])
                    t+=1
                else:
                    thisrow.append("")
            self.table.insert("", 'end', text=str(i), values=thisrow)

        for col in self.table['columns']:
            if col not in headers:
                self.table.column(col, width=0, stretch=False)
    def spass(self):
        pass


Gui()